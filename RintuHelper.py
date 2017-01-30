from openpyxl import Workbook
from openpyxl import load_workbook

########################################
########################################
## MAKE SURE THE FILE IS CLOSED FIRST ##
## MAKE SURE THE FILE IS CLOSED FIRST ##
## MAKE SURE THE FILE IS CLOSED FIRST ##
########################################
########################################

#Enter the file name and the extension here:
FILENAME = input("Enter the file name and the extension here: ")


wb = load_workbook(FILENAME)
ws = wb.active

PERIOD_COLUMN_LETTERS = input("The Period Column letters: ")
ACTIVELEV_COLUMN_LETTERS = input("The ACTIVELEV column letters: ")
COLUMN_TO_WRITE_TO = input("The column that you would like to write info to: ")
ROW_TO_START_ON = input("This number should be the first row number of the info: ")
ROW_TO_END_ON = '100000'



ACTIVELEV = ws[ACTIVELEV_COLUMN_LETTERS+ROW_TO_START_ON:ACTIVELEV_COLUMN_LETTERS+ROW_TO_END_ON]
period_range = ws[PERIOD_COLUMN_LETTERS+ROW_TO_START_ON:PERIOD_COLUMN_LETTERS+ROW_TO_END_ON]
write_range = ws[COLUMN_TO_WRITE_TO+ROW_TO_START_ON:COLUMN_TO_WRITE_TO+ROW_TO_END_ON]
current = ws[PERIOD_COLUMN_LETTERS+ROW_TO_START_ON].value

counter = 0
sums = 0
average = 0.0
items = []

for period, active, write in zip(period_range, ACTIVELEV, write_range):
    if period[0].value == current:
        if active[0].value == 'W':
            sums += 3
        elif active[0].value == 'S':
            sums += 1.5
        elif active[0].value == 'V':
            sums += 6
        else:
            sums += 0
        counter += 1
        items.append(write)
        
    else:
        average = sums / counter
        current = period[0].value
        counter = 1
        sums = 0
        
        if active[0].value == 'W':
            sums += 3
        elif active[0].value == 'S':
            sums += 1.5
        elif active[0].value == 'V':
            sums += 6
        else:
            sums += 0
        
        for cell in items:
            print(cell[0])
            cell[0].value = average
            
        items.clear()
        items.append(write)
        
wb.save(FILENAME)

